import re
import fitz
import pdfplumber
from lxml import etree
import zipfile
import markdown
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from .utils import is_likely_heading, infer_heading_level, is_table_header

import logging

# 基本配置
logging.basicConfig(level=logging.INFO)
log = logging.info  # 或者使用 logger
def clean_text(text):
    """
    Clean text using regex substitutions and count corrections.
    This is extracted from the _clean_text method of PageTextPreparation class.
    """
    command_mapping = {
        'zero': '0',
        'one': '1',
        'two': '2',
        'three': '3',
        'four': '4',
        'five': '5',
        'six': '6',
        'seven': '7',
        'eight': '8',
        'nine': '9',
        'period': '.',
        'comma': ',',
        'colon': ":",
        'hyphen': "-",
        'percent': '%',
        'dollar': '$',
        'space': ' ',
        'plus': '+',
        'minus': '-',
        'slash': '/',
        'asterisk': '*',
        'lparen': '(',
        'rparen': ')',
        'parenright': ')',
        'parenleft': '(',
        'wedge.1_E': '',
    }

    recognized_commands = "|".join(command_mapping.keys())
    slash_command_pattern = rf"/({recognized_commands})(\.pl\.tnum|\.tnum\.pl|\.pl|\.tnum|\.case|\.sups)"

    occurrences_amount = len(re.findall(slash_command_pattern, text))
    occurrences_amount += len(re.findall(r'glyph<[^>]*>', text))
    occurrences_amount += len(re.findall(r'/([A-Z])\.cap', text))

    corrections = []

    def replace_command(match):
        base_command = match.group(1)
        replacement = command_mapping.get(base_command)
        if replacement is not None:
            corrections.append((match.group(0), replacement))
        return replacement if replacement is not None else match.group(0)

    def replace_glyph(match):
        corrections.append((match.group(0), ''))
        return ''

    def replace_cap(match):
        original = match.group(0)
        replacement = match.group(1)
        corrections.append((original, replacement))
        return replacement

    text = re.sub(slash_command_pattern, replace_command, text)
    text = re.sub(r'glyph<[^>]*>', replace_glyph, text)
    text = re.sub(r'/([A-Z])\.cap', replace_cap, text)

    return text, occurrences_amount, corrections


def extract_outline_from_markdown(md_path):
    """
    解析 Markdown 文件，按原顺序提取大纲信息和内容，与docx解析器格式对齐
    """
    try:
        with open(md_path, 'r', encoding='utf-8') as f:
            md_content = f.read()

        # 将 Markdown 转换为 HTML
        html_content = markdown.markdown(md_content, extensions=['extra'])

        # 使用 BeautifulSoup 解析 HTML
        soup = BeautifulSoup(html_content, 'html.parser')

        extracted_content = "Outline and Text:\n"

        # 获取所有需要处理的元素，保持原文档顺序
        elements = soup.find_all(['h1', 'h2', 'h3', 'h4', 'h5', 'h6', 'p', 'ul', 'ol', 'blockquote', 'pre', 'table'])

        # 用于合并连续的非标题内容
        current_paragraph_parts = []
        
        def flush_paragraph():
            """将累积的段落内容输出"""
            nonlocal current_paragraph_parts, extracted_content
            if current_paragraph_parts:
                combined_text = " ".join(current_paragraph_parts).strip()
                if combined_text:
                    extracted_content += f"Paragraph: {combined_text}\n"
                current_paragraph_parts = []

        # 按原文档顺序处理每个元素
        for element in elements:
            if element.name in ['h1', 'h2', 'h3', 'h4', 'h5', 'h6']:
                # 遇到标题时，先输出之前累积的段落内容
                flush_paragraph()
                
                # 处理标题 - 与docx解析器保持一致的格式
                level = int(element.name[1])  # 获取标题级别
                title = element.get_text().strip()
                indent = "  " * (level - 1)
                extracted_content += f"{indent}Heading {level}: {title}\n"

            elif element.name == 'p':
                # 处理段落 - 累积到当前段落中
                text = element.get_text().strip()
                if text:
                    current_paragraph_parts.append(text)

            elif element.name in ['ul', 'ol']:
                # 处理列表 - 累积到当前段落中
                for li in element.find_all('li'):
                    text = li.get_text().strip()
                    if text:
                        current_paragraph_parts.append(text)

            elif element.name == 'blockquote':
                # 处理块引用 - 累积到当前段落中
                text = element.get_text().strip()
                if text:
                    current_paragraph_parts.append(text)

            elif element.name == 'pre':
                # 处理代码块 - 累积到当前段落中
                code = element.find('code')
                if code:
                    text = code.get_text().strip()
                    current_paragraph_parts.append(text)

            elif element.name == 'table':
                # 遇到表格时，先输出之前累积的段落内容
                flush_paragraph()
                
                # 处理表格 - 与docx解析器格式对齐
                extracted_content += f"Table:\n"
                rows = element.find_all('tr')
                for row in rows:
                    cells = row.find_all(['th', 'td'])
                    row_data = []
                    for cell in cells:
                        cell_text = cell.get_text().strip()
                        row_data.append(cell_text)
                    if row_data:
                        extracted_content += "Paragraph: " + " | ".join(row_data) + "\n"

        # 处理完所有元素后，输出最后累积的段落内容
        flush_paragraph()

        return extracted_content

    except Exception as e:
        raise Exception(f"Error parsing Markdown file: {e}")



def extract_outline_from_excel(xlsx_path):
    """
    解析 Excel 文件，提取工作表和單元格內容
    """
    try:
        workbook = load_workbook(xlsx_path)
        extracted_content = "Outline and Text:\n"

        # 處理每個工作表
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

            # 添加工作表作為一級標題
            extracted_content += f"Heading 1: sheet - {sheet_name}\n"

            # 提取單元格內容
            for row in sheet.iter_rows(values_only=True):
                row_data = []
                for cell in row:
                    if cell is not None:
                        row_data.append(str(cell))

                if row_data:
                    # 檢查是否可能是表頭
                    if is_table_header(row_data):
                        extracted_content += f"Heading 2: line\n"

                    # 添加行內容
                    row_text = " | ".join(row_data)
                    extracted_content += f"Paragraph: {row_text}\n"

        return extracted_content

    except Exception as e:
        raise Exception(f"Error parsing Excel file: {e}")


try:
    from langchain.text_splitter import RecursiveCharacterTextSplitter
except ImportError:
    log("Please install langchain: pip install langchain")
    raise


def format_table_as_markdown(table):
    """
    将 pdfplumber 提取的表格 (列表的列表) 格式化为 Markdown 格式。
    """
    markdown_table = ""
    if not table:
        return ""

    # 清理和处理表头
    header = [str(cell).strip() if cell is not None else "" for cell in table[0]]
    markdown_table += "| " + " | ".join(header) + " |\n"

    # 创建 Markdown 分隔符行
    separator = ["---" for _ in header]
    markdown_table += "| " + " | ".join(separator) + " |\n"

    # 处理表格数据行
    for row in table[1:]:
        # 确保行不为空或仅包含空单元格
        if row and any(cell is not None and str(cell).strip() for cell in row):
            cleaned_row = [str(cell).strip() if cell is not None else "" for cell in row]
            markdown_table += "| " + " | ".join(cleaned_row) + " |\n"

    return markdown_table

    
def extract_outline_from_pdf(pdf_path):
    """
    解析 PDF 文件，提取大纲和内容，并将表格格式化为 Markdown。
    """
    try:
        # 使用 PyMuPDF 提取大纲 (TOC)
        doc = fitz.open(pdf_path)
        outline = doc.get_toc()

        # 使用 pdfplumber 提取文本和表格
        with pdfplumber.open(pdf_path) as pdf:
            extracted_content = "Outline and Text:\n"

            # 建立页码到标题的映射
            page_to_headings = {}
            if outline:
                for level, title, page_num in outline:
                    if page_num not in page_to_headings:
                        page_to_headings[page_num] = []
                    page_to_headings[page_num].append((level, title))

            # 逐页处理内容
            for i, page in enumerate(pdf.pages):
                current_page_num = i + 1

                # 如果当前页有标题，先输出标题
                if current_page_num in page_to_headings:
                    for level, title in page_to_headings[current_page_num]:
                        indent = "  " * (level - 1)
                        extracted_content += f"\n{indent}# Heading {level}: {title}\n"

                # 提取并处理表格
                # 调整 table_settings 可以优化对不同格式表格的识别效果
                tables = page.extract_tables(table_settings={
                    "vertical_strategy": "lines",
                    "horizontal_strategy": "lines"
                })

                table_texts_for_removal = []
                if tables:
                    for table_idx, table in enumerate(tables):
                        extracted_content += f"\n--- Table {table_idx + 1} (Page {current_page_num}) ---\n"
                        extracted_content += format_table_as_markdown(table)
                        # 收集表格中的文本，以便从页面总文本中剔除，避免重复
                        for row in table:
                            for cell in row:
                                if cell:
                                    table_texts_for_removal.append(str(cell).strip())

                # 提取页面文本
                text = page.extract_text()
                if text:
                    # 剔除已在表格中处理过的文本
                    for table_text in set(table_texts_for_removal):
                        if table_text:
                            text = text.replace(table_text, "")

                    # 按段落分割
                    paragraphs = text.split('\n\n')
                    for paragraph in paragraphs:
                        cleaned_paragraph = ' '.join(paragraph.split()).strip()
                        if cleaned_paragraph:
                            cleaned_text, _, _ = clean_text(cleaned_paragraph)
                            extracted_content += f"\nParagraph: {cleaned_text}\n"

        doc.close()
        return extracted_content

    except Exception as e:
        raise Exception(f"Error parsing PDF file: {e}")


NS = {'w': 'http://schemas.openxmlformats.org/wordprocessingml/2006/main'}

def parse_xml(docx_zip, path):
    try:
        return etree.fromstring(docx_zip.read(path))
    except KeyError:
        return None

def build_style_outline_map(styles_tree):
    """
    返回 {styleId: level}，level 从 1 开始
    优先级：
    1) 样式中显式 w:outlineLvl
    2) 样式名匹配 Heading1/Heading2 ...
    """
    mapping = {}
    if styles_tree is None:
        return mapping
    for st in styles_tree.xpath('//w:style', namespaces=NS):
        style_id = st.get(f'{{{NS["w"]}}}styleId')
        lvl = st.xpath('.//w:outlineLvl', namespaces=NS)
        if lvl:
            try:
                mapping[style_id] = int(lvl[0].get(f'{{{NS["w"]}}}val')) + 1
                continue
            except (TypeError, ValueError):
                pass
        if style_id and style_id.lower().startswith('heading'):
            m = re.search(r'(\d+)', style_id)
            if m:
                mapping[style_id] = int(m.group(1))
    return mapping

def get_paragraph_text(p):
    runs = p.xpath('.//w:r//w:t | .//w:fldSimple//w:t | .//w:hyperlink//w:t', namespaces=NS)
    return ''.join((t.text or '') for t in runs).strip()

def get_paragraph_style_id(p):
    ps = p.xpath('.//w:pPr/w:pStyle', namespaces=NS)
    if ps:
        return ps[0].get(f'{{{NS["w"]}}}val')
    return None

def get_paragraph_outline_level(p):
    lvl = p.xpath('.//w:pPr/w:outlineLvl', namespaces=NS)
    if lvl:
        try:
            return int(lvl[0].get(f'{{{NS["w"]}}}val')) + 1
        except (TypeError, ValueError):
            return None
    return None

def get_list_level(p):
    ilvl = p.xpath('.//w:pPr/w:numPr/w:ilvl', namespaces=NS)
    if ilvl:
        try:
            return int(ilvl[0].get(f'{{{NS["w"]}}}val'))
        except (TypeError, ValueError):
            pass
    return None

def detect_heading_level_by_text(text, max_level=6):
    """
    兜底匹配：形如
    - "1 引言", "1. 引言"
    - "1.1 定义", "3.2.1 内容"
    返回分段数作为层级，否则 None
    """
    m = re.match(r'^\s*(\d+(?:\.\d+){0,10})[\.\s]+', text)
    if not m:
        return None
    parts = m.group(1).split('.')
    return min(len(parts), max_level)

def extract_table(tbl):
    rows_data = []
    rows = tbl.xpath('./w:tr', namespaces=NS)
    for r in rows:
        cells = r.xpath('./w:tc', namespaces=NS)
        row_texts = []
        for c in cells:
            t = ''.join((t.text or '') for t in c.xpath('.//w:t', namespaces=NS)).strip()
            row_texts.append(t)
        # 保留空单元格，但去掉整行全空
        if any(x.strip() for x in row_texts):
            rows_data.append(' | '.join(row_texts))
    if rows_data:
        return "Table:\n" + '\n'.join(rows_data)
    return None

def extract_outline_from_docx_with_tables(docx_path):
    """
    以 XML 解析为主的 docx 大纲/文本/表格抽取：
    - 标题判定优先级：pStyle -> 段落 outlineLvl -> 数字分级文本兜底
    - 列表只用于缩进显示（不当成标题）
    - 连续普通段落合并，遇到标题/表格/空行 flush
    """
    try:
        with zipfile.ZipFile(docx_path, 'r') as z:
            doc = parse_xml(z, 'word/document.xml')
            styles = parse_xml(z, 'word/styles.xml')
            # numbering 保留以便扩展（当前仅用 ilvl 缩进）
            _ = parse_xml(z, 'word/numbering.xml')

        if doc is None:
            raise RuntimeError('word/document.xml not found')

        style_map = build_style_outline_map(styles)

        extracted = []
        para_group = []

        def flush_group():
            nonlocal para_group
            if para_group:
                extracted.append("Paragraph: " + '\n'.join(para_group))
                para_group = []

        body_elems = doc.xpath('//w:body/*', namespaces=NS)

        for el in body_elems:
            tag_local = etree.QName(el.tag).localname

            if tag_local == 'p':
                text = get_paragraph_text(el)
                if not text:
                    # 空段落作为段落边界
                    flush_group()
                    continue

                # 标题判定
                style_id = get_paragraph_style_id(el)
                heading_level = None

                if style_id and style_id in style_map:
                    heading_level = style_map[style_id]
                else:
                    heading_level = get_paragraph_outline_level(el)
                    if heading_level is None:
                        heading_level = detect_heading_level_by_text(text)

                if heading_level:
                    flush_group()
                    indent = '  ' * (heading_level - 1)
                    extracted.append(f"{indent}Heading {heading_level}: {text}")
                else:
                    ilvl = get_list_level(el)
                    if ilvl is not None:
                        indent = '  ' * ilvl
                        para_group.append(f"{indent}- {text}")
                    else:
                        para_group.append(text)

            elif tag_local == 'tbl':
                flush_group()
                t = extract_table(el)
                if t:
                    extracted.append(t)

            else:
                # 其他元素直接作为边界
                flush_group()

        flush_group()
        return "Outline and Text:\n" + '\n'.join(extracted)

    except Exception as e:
        raise Exception(f"Error parsing docx file: {e}")
